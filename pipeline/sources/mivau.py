"""
Source MIVAU — Estadística de Precios de Vivienda Libre.

El portal del Ministerio (mivau.gob.es) bloquea peticiones automatizadas
con HTTP 403. Estrategia pragmática:

1. El usuario descarga manualmente el Excel trimestral desde:
   https://www.mivau.gob.es/vivienda/datos-y-estadisticas
   → "Estadística de precios de vivienda" → Excel trimestre actual
2. Lo guarda en data/raw/mivau_YYYYQN.xlsx
3. Este parser lo lee y devuelve {nombre_norm: precio_m2}

Cobertura habitual: capitales + municipios > 25.000 habitantes (~250 de 329).
"""
import re
import unicodedata
from pathlib import Path
from openpyxl import load_workbook
from ..config import RAW_DIR

def _normalize(s: str) -> str:
    s = unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", s).lower().strip()

def find_mivau_file() -> Path | None:
    """Busca el archivo más reciente mivau_*.xlsx en data/raw/."""
    candidates = sorted(RAW_DIR.glob("mivau_*.xlsx"), reverse=True)
    if not candidates:
        candidates = sorted(RAW_DIR.glob("*precios*vivienda*.xlsx"), reverse=True)
    return candidates[0] if candidates else None

def parse_mivau(xlsx_path: Path) -> dict:
    """
    Parsea Excel MIVAU y devuelve {nombre_normalizado: precio_eur_m2}.

    El Excel típico tiene columnas: 'Provincia', 'Municipio', 'Trimestre N',
    'Precio €/m²' o similar. Heurística para detectar la estructura.
    """
    if not xlsx_path or not xlsx_path.exists():
        print(f"[MIVAU] archivo no encontrado: {xlsx_path}")
        return {}
    print(f"[MIVAU] parseando {xlsx_path.name}...")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    out = {}
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows: continue
        # Buscar fila header
        header_idx = None
        for i, r in enumerate(rows[:15]):
            if r and any(c and "MUNICIPIO" in str(c).upper() for c in r):
                header_idx = i
                break
        if header_idx is None: continue
        header = [str(c).upper().strip() if c else "" for c in rows[header_idx]]
        col_muni = next((i for i,h in enumerate(header) if "MUNICIPIO" in h), None)
        # Buscar columna de precio (la última numérica suele ser la actual)
        col_price = None
        for i, h in enumerate(header):
            if "PRECIO" in h or "€/M" in h or "EUROS/M" in h:
                col_price = i
        if col_muni is None or col_price is None: continue
        for r in rows[header_idx+1:]:
            if not r or r[col_muni] is None: continue
            muni = str(r[col_muni]).strip()
            try:
                p = float(str(r[col_price]).replace(",", "."))
                if 200 < p < 12000:  # rango razonable €/m²
                    out[_normalize(muni)] = round(p)
            except (ValueError, TypeError):
                continue
    print(f"[MIVAU] extraídos {len(out)} municipios")
    return out

def fetch_prices() -> dict:
    """API pública del módulo."""
    fp = find_mivau_file()
    return parse_mivau(fp) if fp else {}

def match_to_cities(prices: dict, cities: list) -> dict:
    """Match {slug: precio_m2}."""
    result = {}
    for c in cities:
        norm = _normalize(c["nombre"])
        if norm in prices:
            result[c["slug"]] = prices[norm]
    print(f"[MIVAU] match: {len(result)}/{len(cities)}")
    return result

if __name__ == "__main__":
    p = fetch_prices()
    print(f"Total municipios MIVAU: {len(p)}")
    if p:
        print("Sample:", list(p.items())[:5])
    else:
        print(f"\nINSTRUCCIONES MANUALES:")
        print(f"1. Visita https://www.mivau.gob.es/vivienda/datos-y-estadisticas")
        print(f"2. Descarga el Excel trimestral 'Precios de Vivienda Libre'")
        print(f"3. Guárdalo en {RAW_DIR}/mivau_2026Q1.xlsx")
        print(f"4. Vuelve a ejecutar este módulo")
