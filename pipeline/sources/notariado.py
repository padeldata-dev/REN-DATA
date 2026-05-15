"""
Source Centro de Información Estadística Notarial (CIEN).

El Consejo General del Notariado publica trimestralmente la
'Estadística Notarial' en PDF y Excel. Cobertura: provincias y
capitales (no municipios pequeños).

Como con MIVAU, descarga manual a data/raw/notariado_YYYYQN.xlsx.
Aporta validación cruzada del precio €/m² escriturado y nº transacciones.
"""
import re
import unicodedata
from pathlib import Path
from openpyxl import load_workbook
from ..config import RAW_DIR

def _normalize(s: str) -> str:
    s = unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", s).lower().strip()

def find_file() -> Path | None:
    candidates = sorted(RAW_DIR.glob("notariado_*.xlsx"), reverse=True)
    return candidates[0] if candidates else None

def parse(xlsx_path: Path) -> dict:
    """Devuelve {nombre_norm: {'precio': X, 'transacciones': Y}}."""
    if not xlsx_path or not xlsx_path.exists():
        return {}
    print(f"[NOTARIADO] parseando {xlsx_path.name}...")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    out = {}
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows: continue
        header_idx = next((i for i,r in enumerate(rows[:10])
                          if r and any(c and "PROVINCIA" in str(c).upper() for c in r)), None)
        if header_idx is None: continue
        header = [str(c).upper().strip() if c else "" for c in rows[header_idx]]
        col_p = next((i for i,h in enumerate(header) if "PROVINCIA" in h), None)
        col_pr = next((i for i,h in enumerate(header) if "PRECIO" in h), None)
        col_tx = next((i for i,h in enumerate(header) if "TRANS" in h or "OPERAC" in h), None)
        if col_p is None: continue
        for r in rows[header_idx+1:]:
            if not r or r[col_p] is None: continue
            name = _normalize(str(r[col_p]))
            entry = {}
            if col_pr is not None and r[col_pr] is not None:
                try: entry["precio"] = round(float(str(r[col_pr]).replace(",",".")))
                except: pass
            if col_tx is not None and r[col_tx] is not None:
                try: entry["transacciones"] = int(r[col_tx])
                except: pass
            if entry:
                out[name] = entry
    print(f"[NOTARIADO] extraídos {len(out)} ámbitos")
    return out

def fetch():
    fp = find_file()
    return parse(fp) if fp else {}

if __name__ == "__main__":
    d = fetch()
    print(f"Total: {len(d)}")
    if not d:
        print(f"\nINSTRUCCIONES MANUALES:")
        print(f"1. Visita https://www.notariado.org/portal/centro-de-informacion-estadistica-notarial")
        print(f"2. Descarga el Excel trimestral de 'Estadística Notarial — Vivienda'")
        print(f"3. Guárdalo en {RAW_DIR}/notariado_2026Q1.xlsx")
