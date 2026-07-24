"""
Source INE — población por municipio.

Estrategia: descargar pobmun.zip (zip oficial con XLS por año/provincia)
y construir un mapping {nombre_normalizado: poblacion}. Match contra
el master por nombre normalizado (sin acentos, lowercase).

API alternativa: https://servicios.ine.es/wstempus/js/ES/DATOS_TABLA/{tabla_id}
con tabla_id por provincia (2855..2903). Más complejo, mismo resultado.
"""
import io
import zipfile
import unicodedata
from pathlib import Path
import requests
from openpyxl import load_workbook
from ..config import RAW_DIR, USER_AGENT

POBMUN_URL = "https://www.ine.es/pob_xls/pobmun.zip"
POBMUN_FALLBACK_URL = "https://www.ine.es/pob_xls/pobmun24.xls"  # último año típico

def _normalize(s: str) -> str:
    """Quita acentos, lowercase, strip."""
    s = unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode("ascii")
    return s.lower().strip()

def _slug_to_norm(slug: str) -> str:
    """Convierte 'san-sebastian-de-los-reyes' → 'san sebastian de los reyes'."""
    # Quitar sufijos provinciales del slug (ej: 'san-fernando-cadiz' → 'san fernando')
    base = slug.replace("-", " ")
    return base

# Artículos co-oficiales que el INE antepone invertidos ("Porriño, O").
# Incluye castellano, catalán/valenciano, gallego y balear.
COOFICIAL_ARTICLES = ("la", "el", "los", "las", "les", "els", "l'", "o", "a", "os", "as", "es", "sa")

def _deinvert_article(name: str) -> str:
    """
    Invierte la forma INE "X, Art" → "Art X" de forma genérica.
      'porrino, o'              → 'o porrino'
      'franqueses del valles, les' → 'les franqueses del valles'
      "alfas del pi, l'"        → "l'alfas del pi"   (l' sin espacio)
    Si no hay artículo antepuesto reconocido, devuelve el nombre sin tocar.
    """
    if "," not in name:
        return name
    base, art = name.rsplit(",", 1)
    base, art = base.strip(), art.strip()
    if art in COOFICIAL_ARTICLES:
        return f"{art}{base}" if art.endswith("'") else f"{art} {base}"
    return name

def _expand_ine_keys(populations: dict) -> dict:
    """
    Índice ampliado de poblaciones para casar nombres-web con claves INE.
    Para cada clave INE añade (sin sobrescribir las originales):
      - cada parte de las formas bilingües separadas por '/'
        ('montcada/moncada' → 'montcada', 'moncada')
      - la forma de-invertida de cada parte
        ('vila joiosa, la/villajoyosa' → 'la vila joiosa', 'villajoyosa')
    """
    idx = dict(populations)
    for k, v in populations.items():
        for part in (k.split("/") if "/" in k else [k]):
            part = part.strip()
            if not part:
                continue
            idx.setdefault(part, v)
            idx.setdefault(_deinvert_article(part), v)
    return idx

def download_pobmun(force: bool = False) -> Path:
    """Descarga pobmun.zip al cache RAW. Retorna path al ZIP."""
    out = RAW_DIR / "pobmun.zip"
    if out.exists() and not force:
        return out
    print(f"[INE] descargando {POBMUN_URL} ...")
    r = requests.get(POBMUN_URL, headers={"User-Agent": USER_AGENT}, timeout=60)
    r.raise_for_status()
    out.write_bytes(r.content)
    print(f"[INE] guardado en {out} ({len(r.content)/1024:.0f} KB)")
    return out

def parse_pobmun(zip_path: Path) -> dict:
    """
    Extrae {nombre_normalizado_municipio: poblacion_total}.
    El zip contiene XLS/XLSX con datos por provincia/municipio.
    """
    populations = {}
    with zipfile.ZipFile(zip_path) as zf:
        # Ordenar por AÑO ascendente para que el más reciente se procese el
        # último y sus valores prevalezcan sobre años anteriores en claves que
        # colisionen (evita que pobmun96.xlsx pise a pobmun25.xlsx por orden
        # alfabético).
        import re as _re
        def _year_of(n: str) -> int:
            m = _re.search(r"pobmun(\d{2})", n.lower())
            if not m:
                return -1
            yy = int(m.group(1))
            return 2000 + yy if yy <= 50 else 1900 + yy
        names = sorted([n for n in zf.namelist() if n.lower().endswith((".xls", ".xlsx"))],
                       key=_year_of)
        if not names:
            print("[INE] AVISO: el ZIP no contiene XLS/XLSX")
            return populations
        # Procesar TODOS los xlsx; los xls puros (formato antiguo) se omiten silenciosamente
        for name in names:
            if not name.lower().endswith(".xlsx"):
                continue
            try:
                with zf.open(name) as fh:
                    data = io.BytesIO(fh.read())
                wb = load_workbook(data, read_only=True, data_only=True)
                ws = wb.active
                # Buscar columnas: típicamente CPRO, CMUN, NOMBRE, POB_TOT
                # Heurística: la fila con esos textos es el header
                rows = ws.iter_rows(values_only=True)
                header = None
                for i, row in enumerate(rows):
                    if i > 5: break
                    if row and any(c and "NOMBRE" in str(c).upper() for c in row):
                        header = [str(c).upper().strip() if c else "" for c in row]
                        break
                if not header:
                    continue
                col_name = next((i for i,h in enumerate(header) if h == "NOMBRE"), None)
                # Buscar columna de población total
                col_pop = None
                for i, h in enumerate(header):
                    if "TOTAL" in h or "POB" in h:
                        col_pop = i
                        break
                if col_name is None or col_pop is None:
                    continue
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or row[col_name] is None:
                        continue
                    nombre = str(row[col_name]).strip()
                    pop_raw = row[col_pop]
                    try:
                        pop = int(pop_raw)
                    except (ValueError, TypeError):
                        continue
                    populations[_normalize(nombre)] = pop
            except Exception as e:
                print(f"[INE] WARN parsing {name}: {e}")
                continue
    return populations

def fetch_populations(force: bool = False) -> dict:
    """API pública: devuelve {nombre_normalizado: poblacion}."""
    zp = download_pobmun(force=force)
    return parse_pobmun(zp)

def match_to_cities(populations: dict, cities: list) -> dict:
    """
    Devuelve {slug: poblacion} matcheando por nombre normalizado.
    Heurística para slugs con sufijo provincial + aliases co-oficiales.
    """
    # Aliases para nombres co-oficiales y casos especiales
    ALIASES = {
        "castellon": "castello de la plana", "castellon-de-la-plana": "castello de la plana",
        "alcoy": "alcoi", "elche": "elche/elx", "sagunto": "sagunt/sagunto",
        "orense": "ourense", "la-coruna": "coruna, a", "a-coruna": "coruna, a",
        "vitoria-gasteiz": "vitoria-gasteiz", "san-sebastian-donostia": "donostia/san sebastian",
        "donostia": "donostia/san sebastian", "san-sebastian": "donostia/san sebastian",
        "las-palmas-gc": "palmas de gran canaria, las", "fuerteventura-pjto-rosario": "puerto del rosario",
        "lanzarote-arrecife": "arrecife", "ibiza": "eivissa", "mahon": "mao",
        "vigo-pontevedra": "vigo", "san-fernando-cadiz": "san fernando",
        "javea": "xabia/javea", "denia": "denia", "calpe": "calp",
        "ondara": "ondara", "benisa": "benissa",
        "esplugues-de-llobregat": "esplugues de llobregat", "viladecans": "viladecans",
        "sant-cugat-del-valles": "sant cugat del valles",
        "cerdanyola-del-valles": "cerdanyola del valles",
        "premia-de-mar": "premia de mar", "el-prat-de-llobregat": "prat de llobregat, el",
        "el-puerto-de-santa-maria": "puerto de santa maria, el",
        "la-laguna": "san cristobal de la laguna", "san-cristobal-de-la-laguna": "san cristobal de la laguna",
        "san-vicente-del-raspeig": "san vicent del raspeig/sant vicent del raspeig",
        "vila-real": "vila-real", "vila-seca": "vila-seca", "soller": "soller",
        "felanitx": "felanitx", "marratxi": "marratxi", "calvia": "calvia",
        "manacor": "manacor", "alcudia": "alcudia",
        "ciutadella-de-menorca": "ciutadella de menorca",
        "navalmoral-de-la-mata": "navalmoral de la mata",
        "tossa-de-mar": "tossa de mar", "lloret-de-mar": "lloret de mar",
        "blanes": "blanes", "calella": "calella", "calafell": "calafell",
        "vilanova-i-la-geltru": "vilanova i la geltru",
        "santa-cruz-de-tenerife": "santa cruz de tenerife",
        "santa-cruz-de-la-palma": "santa cruz de la palma",
        "san-bartolome-de-tirajana": "san bartolome de tirajana",
        "santa-lucia-de-tirajana": "santa lucia de tirajana",
        "ciudad-real": "ciudad real", "ciudad-rodrigo": "ciudad rodrigo",
        "alcala-la-real": "alcala la real", "guardamar-del-segura": "guardamar del segura",
        "pilar-de-la-horadada": "pilar de la horadada",
        "moron-de-la-frontera": "moron de la frontera",
        "chiclana-de-la-frontera": "chiclana de la frontera",
        "conil-de-la-frontera": "conil de la frontera",
        "san-fernando-de-henares": "san fernando de henares",
        "san-lorenzo-de-el-escorial": "san lorenzo de el escorial",
        "san-sebastian-de-los-reyes": "san sebastian de los reyes",
        "torrejon-de-ardoz": "torrejon de ardoz",
        "boadilla-del-monte": "boadilla del monte",
        "rivas-vaciamadrid": "rivas-vaciamadrid",
        "molina-de-segura": "molina de segura",
        "san-pedro-del-pinatar": "san pedro del pinatar",
        "puerto-de-la-cruz": "puerto de la cruz",
        "los-realejos": "realejos, los",
        "l-hospitalet-de-llobregat": "hospitalet de llobregat, l'",
        "vitoria": "vitoria-gasteiz",
        "las-rozas": "rozas de madrid, las",
        "cangas-do-morrazo": "cangas",
        "l-escala": "escala, l'",
        "mondragon": "arrasate/mondragon",
        "luarca": "valdes",  # Luarca es la capital del concejo de Valdés
        "velez-blanco": "velez-blanco",
        # Formas bilingües / artículo antepuesto invertido del INE (padrón 2025).
        # Con el índice ampliado (_expand_ine_keys) casan solas, pero se dejan
        # explícitas como red de seguridad y documentación.
        "la-vila-joiosa": "vila joiosa, la/villajoyosa",
        "la-vall-d-uixo": "vall d'uixo, la",
        "montcada": "montcada/moncada",
        "l-alfas-del-pi": "alfas del pi, l'",
        "o-porrino": "porrino, o",
        "franqueses-del-valles-les": "franqueses del valles, les",
        "a-estrada": "estrada, a",
    }
    # Índice ampliado: añade formas de-invertidas y partes bilingües de las
    # claves INE, para casar nombres-web como "O Porriño" o "Les Franqueses".
    pop_idx = _expand_ine_keys(populations)
    result = {}
    not_found = []
    for c in cities:
        slug = c["slug"]
        nombre = c["nombre"]
        # 1) Nombre directo
        norm_n = _normalize(nombre)
        if norm_n in populations:
            result[slug] = populations[norm_n]; continue
        # 2) Alias por slug (tabla manual: tiene prioridad sobre la heurística)
        if slug in ALIASES:
            ali = _normalize(ALIASES[slug])
            if ali in populations:
                result[slug] = populations[ali]; continue
        # 2b) Nombre directo contra el índice ampliado (bilingües + de-invertido)
        if norm_n in pop_idx:
            result[slug] = pop_idx[norm_n]; continue
        # 3) Sin sufijo paréntesis
        if "(" in nombre:
            base = _normalize(nombre.split("(")[0])
            if base in populations:
                result[slug] = populations[base]; continue
        # 4) Slug normalizado, eliminar sufijo provincial
        norm_slug = _normalize(_slug_to_norm(slug))
        for suff in (" cadiz", " madrid", " mallorca", " baleares", " gc", " pontevedra", " arrecife"):
            if norm_slug.endswith(suff):
                norm_slug = norm_slug[: -len(suff)].strip()
                break
        if norm_slug in populations:
            result[slug] = populations[norm_slug]; continue
        # 5) Búsqueda flexible: nombre comienza por el slug
        candidates = [k for k in populations if k.startswith(norm_slug + " ") or k.startswith(norm_slug + ",") or k.startswith(norm_slug + "/")]
        if len(candidates) == 1:
            result[slug] = populations[candidates[0]]; continue
        # 6) Inversión por artículo: "la-oliva" → "oliva, la"
        for art in ("la ", "el ", "los ", "las "):
            if norm_slug.startswith(art):
                inv = f"{norm_slug[len(art):]}, {art.strip()}"
                if inv in populations:
                    result[slug] = populations[inv]; break
        else:
            not_found.append(slug); continue
        # bucle break sin entrar al else
        if slug not in result:
            not_found.append(slug)
    if not_found:
        print(f"[INE] {len(not_found)} ciudades sin match: {not_found[:10]}...")
    print(f"[INE] match: {len(result)}/{len(cities)}")
    return result

if __name__ == "__main__":
    pops = fetch_populations()
    print(f"Total municipios INE: {len(pops)}")
    print("Sample:", list(pops.items())[:5])
